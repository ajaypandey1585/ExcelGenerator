import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Create a DataFrame from the provided data
data = {
    'VALORNR': [20129, 20131, 20133, 20135, 20137, 20139, 20141, 20143, 20145, 20147, 20149, 20151, 20153, 20155, 20157, 20159, 20161, 43345, 66422, 89499, 201597, 209159, 158730, 181807, 204884, 227961, 242142, 263011, 283880, 304749, 325618, 346487, 367356, 388225, 409094, 429963, 450832, 471701, 492570, 513439, 534308, 555177, 11220, 12098, 12976, 13854, 14732, 15610, 16488, 17366],
    'BC': [1795, 2018, 2241, 2464, 2687, 2910, 3133, 3356, 3579, 3802, 4025, 4248, 4471, 4694, 4917, 5140, 5363, 111, 112, 113, 114, 115, 116, 117, 118, 119, 224, 212, 200, 188, 176, 164, 152, 140, 128, 116, 104, 92, 80, 68, 56, 44, 907, 786, 665, 544, 423, 302, 181, 60],
    'CC': [1377, 1512, 1647, 1782, 1917, 2052, 2187, 2322, 2457, 2592, 2727, 2862, 2997, 3132, 3267, 3402, 3537, 4211, 221, 400, 879, 1358, 1837, 2316, 2795, 3274, 351, 402, 453, 504, 555, 606, 657, 708, 759, 810, 861, 912, 963, 1014, 1065, 1116, 807, 876, 945, 1014, 1083, 1152, 1221, 1290]
}

df = pd.DataFrame(data)

# Write the DataFrame to an Excel file
output_file = 'ProdExcel.xlsx'
df.to_excel(output_file, index=False)

# Load the workbook
wb = load_workbook(output_file)
ws = wb.active

# Insert a new row at the beginning and merge cells
ws.insert_rows(1)
ws.merge_cells('A1:C1')

# Set the value and style for the merged cell
cell = ws['A1']
cell.value = 'Prod Data'
cell.alignment = Alignment(horizontal='center', vertical='center')

# Save the workbook
wb.save(output_file)

print(f"Excel file '{output_file}' has been created with the 'Prod Data' header.")


# Load the first Excel file
df_dev = pd.read_excel('DevExcel.xlsx')

# Load the second Excel file
df_prod = pd.read_excel('ProdExcel.xlsx')

# Combine the data from both Excel files
df_combined = pd.concat([df_dev, df_prod], ignore_index=True)

# Create a new Excel writer object
output_file = 'Comparison_Output_Excel.xlsx'
with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
    # Write the combined dataframe to the Excel file
    df_combined.to_excel(writer, index=False, startrow=0, startcol=0)

    # Access the workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Merge the cells in the first row
    max_column = df_combined.shape[1] - 1
    worksheet.merge_range(0, 0, 0, max_column, 'Merged Header Row')

print(f"Excel file '{output_file}' has been created with combined data from both input files.")