import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def write_to_excel(df, filename):
    wb = Workbook()
    ws = wb.active
    for r_idx, row in enumerate(df.iterrows(), start=1):
        for c_idx, cell in enumerate(row[1], start=1):
            ws.cell(row=r_idx, column=c_idx, value=cell)
    
    # Apply conditional formatting to highlight changed cells and use '-->' delimiter
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and '-->' in cell.value:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    wb.save(filename)

production = pd.read_excel('output.xlsx')
LMDB = pd.read_excel('output1.xlsx')

comparevalues = production.values == LMDB.values

print(comparevalues)

rows, cols = np.where(comparevalues==False)

for item in zip(rows, cols):
    production.iloc[item[0], item[1]] = ' {} --> {} '.format(production.iloc[item[0], item[1]], LMDB.iloc[item[0], item[1]])

write_to_excel(production, 'result.xlsx')
