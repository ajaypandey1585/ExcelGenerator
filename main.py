import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Create a DataFrame from the provided data
data = {
    'VALORNR': [20115, 20117, 20119, 20121, 20123, 20125, 20127, 20129, 20131, 20133, 20135, 20137, 20139, 20141, 20143, 20145, 20147, 20149, 20151, 20153, 20155, 20157, 20159, 20161, 33452, 54321, 75190, 96059, 116928, 137797, 43345, 179535, 200404, 221273, 242142, 209159, 181807, 304749, 325618, 346487, 367356, 388225, 409094, 429963, 450832, 471701, 492570, 513439, 534308, 555177],
    'BC': [234, 457, 680, 903, 1126, 1349, 1572, 1795, 2018, 2241, 2464, 2687, 2910, 3133, 3356, 3579, 3802, 4025, 4248, 4471, 4694, 4917, 5140, 5363, 344, 332, 320, 308, 296, 284, 272, 260, 248, 236, 224, 212, 117, 188, 176, 164, 152, 140, 128, 116, 104, 92, 80, 68, 56, 44],
    'CC': [432, 567, 702, 837, 972, 1107, 1242, 1377, 1512, 1647, 1782, 1917, 2052, 2187, 2322, 2457, 2592, 2727, 2862, 2997, 3132, 3267, 3402, 3537, 345, 234, 123, 12, 45, 96, 147, 198, 249, 300, 351, 1358, 453, 504, 555, 606, 657, 708, 759, 810, 861, 912, 963, 1014, 1065, 1116]
}

df = pd.DataFrame(data)

# Write the DataFrame to an Excel file
output_file = 'DevExcel.xlsx'
df.to_excel(output_file, index=False)

# Load the workbook
wb = load_workbook(output_file)
ws = wb.active

# Insert a new row at the beginning and merge cells
ws.insert_rows(1)
ws.merge_cells('A1:C1')

# Set the value and style for the merged cell
cell = ws['A1']
cell.value = 'Dev Excel Data'
cell.alignment = Alignment(horizontal='center', vertical='center')

# Save the workbook
wb.save(output_file)

print(f"Excel file '{output_file}' has been created with the 'Dev Excel Data' header.")
